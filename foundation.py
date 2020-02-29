from openpyxl import Workbook
import openpyxl

ALPHABETS=['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

# wb = openpyxl.load_workbook('Book1.xlsx')
# sheet = wb.get_sheet_by_name('Data')
# countyData = {}

# for row in range(2, sheet.max_row + 1):
#        # Each row in the spreadsheet has data for one census tract.
#        state  = sheet['B' + str(row)].value
#        county = sheet['C' + str(row)].value
#        pop    = sheet['D' + str(row)].value
# wb.close()




def read_excel_data(workbook, sheet , cellrange='A1:Z1000'):
    wb = openpyxl.load_workbook(workbook)
    sheet = wb.get_sheet_by_name(sheet)

    column_limit=1
    row_limit=1000
    columns=[]
    dt=[]
    #choose the columns that are really needed
    for items in ALPHABETS:
        if(sheet[items+'1'].value!=None):
            columns.append(sheet[items+'1'].value)
            column_limit=column_limit+1
        else:
            break
    dt.append(columns)
    for row in range(2,sheet.max_row+1):
        row_data=[]
        for c in range(1,column_limit+1):
            row_data.append(sheet[ALPHABETS[c]+str(row)].value)
        dt.append(row_data)
    wb.close()
    return dt

x=read_excel_data('Book1.xlsx','Data')
print(x)